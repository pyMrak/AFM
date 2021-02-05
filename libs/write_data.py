# -*- coding: utf-8 -*-
"""
Created on Thu Jan 12 07:39:29 2017

@author: andmra2
"""


from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import  PatternFill, Color#, Style, Font
#from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import  Reference, ScatterChart, Series #, LineChart
#from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.drawing.image import Image
from traceback import print_exception
from sys import  exc_info

from libs import read_data as rd
from libs import data_eval as de
from libs import data_manipulation as dm
from libs import TC_report_module as TCrm
from libs.logger import messageLogger
from openpyxl.drawing.text import CharacterProperties, Paragraph, ParagraphProperties, RegularTextRun
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.text import RichText
from PIL import Image as PILImage
from io import BytesIO

from libs import globalPaths


#import math




def write_file(gdat):
    ml = messageLogger()
    ml.writeProgressInfo('Zapisujem podatke...', gdat)
    #gdat.columns=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
    #l=len(gdat.columns)
    #le=len(gdat.w_data[0][0])
    #for i in range(l,le):
    #    gdat.columns.append(gdat.columns[math.floor(i/(l))-1]+gdat.columns[i%(l)])
    #print(columns)
    gdat.wb = Workbook()
    gdat.ws = [gdat.wb.active]
    gdat.ws[0].title = "1 do 10"
    for i in range(len(gdat.w_data)):
        if i>0:
            gdat.ws.append(gdat.wb.create_sheet(str(i*10+1)+" do "+str(((i+1)*10))))
        
        for j in range(len(gdat.w_data[i])):
            for k in range(len(gdat.w_data[i][j])):
#                if k>l-1:
#                    kn=columns[math.floor(k/(l))-1]+columns[k%(l)]
#                else:
#                    kn=columns[k]
                try:
                    gdat.ws[i][gdat.columns[k]+str(j+2)]=gdat.w_data[i][j][k]
                except:
                    pass
     
    for j in range(len(gdat.ws)):  
        l1=0
        for i in range(10):
            l=len(gdat.data[j][i]) 
            try:
                gdat.ws[j].merge_cells(gdat.columns[l1]+'1:'+gdat.columns[l1+l-1]+'1')
                gdat.ws[j][gdat.columns[l1]+'1']=j*10+i+1
            except:
                break
            l1+=l
        gdat.ws[j].freeze_panes = gdat.ws[j]['A3']
        
    ml.writeProgressInfo('Podatki zapisani.', gdat)
    

    gdat=write_table(gdat)
    
    return save(gdat)
    
    
def saveDefFolder(folder):
       file = open('defaultFolder.txt', 'w', encoding="utf-8")
       #print(folder)
       file.write(folder)
#       file.save()
       file.close()
       
       
def saveSettings(nr, newLine, username):
    file = open(globalPaths.path.settings+username+'.set', 'a', encoding="utf-8")
    #lines = file.readlines()
    #file.close()
    #lines[nr] = newLine + '\n'
    file.write(newLine + '\n')
#    file = open(globalPaths.path.settings+username+'.set', 'w')
#    for line in lines:
#        file.write(line)
    file.close()
    
    
def save(gdat,file='xlsx', file_name="AFM.xlsx"):
    ml = messageLogger()
    ml.writeProgressInfo('Shranjujem... ', gdat)
    if file=='xlsx':
        file = gdat.wb
#    elif file=='txt':
#        file = open(file_name + '.txt', 'w')
    if 'measurements' in gdat.folder:
        path = ''
        label = ''
    else:
        path = gdat.folder + '/'
        label = gdat.folder.split('/')[-1].split('\\')[-1]

    if 'Instrumented glow plug' in file_name:
        label = ''

#    kon = ''
    for i in range(2):
        try:
            if i == 0:
                #print('l:', file)
                file.save(path + label + file_name)#.strip('.xlsx')+str(kon)+'.xlsx')
            else:
                #print('m:', file)
                file.save(path + label + file_name.strip('.xlsx')+'('+str(i)+').xlsx')
                file_name = file_name.strip('.xlsx')+'('+str(i)+').xlsx'
            ml.writeProgressInfo('Datoteka shranjena kot '+label+file_name, gdat)
            break
        except Exception as e:
            if i == 1:
                ml.writeError('Shranjevanje neuspešno.', gdat)
                exc_type, exc_value, exc_traceback = exc_info()
                print_exception(exc_type, exc_value, exc_traceback, file='log.log')
#    try:
#        print('Saving data in ' + path + label + file_name.strip('.xlsx')+str(kon)+'.xlsx...')
#        file.save(path + label + file_name.strip('.xlsx')+str(kon)+'.xlsx')
#    except:
#        print('File with same name opened, changing name..')
#        if kon == '':
#            kon=0
#        else:
#            kon += 1
    
    return path + label + file_name
    
def write_QC(gdat):
    if gdat.TC:
        gdat.table += 2
    if gdat.R:
        gdat.table += 1
    if gdat.gdat.nastavitveA.TT:
        gdat.table += 1
    #print('pišem QC')
    gdat.ws[0][gdat.columns[gdat.table]+'2']='QC'
    temp=gdat.QC.name.split('-')
    gdat.ws[0][gdat.columns[gdat.table]+'3']=temp[0]
    try:
        gdat.ws[0][gdat.columns[gdat.table]+'4']=temp[1]+' '+temp[2]
    except:
        pass
    gdat.ws[0][gdat.columns[gdat.table+1]+'2']='t'
    gdat.ws[0].merge_cells(gdat.columns[gdat.table+2]+'2:'+gdat.columns[gdat.table+3]+'2')
    gdat.ws[0][gdat.columns[gdat.table+2]+'2']=gdat.QC.TtR
    gdat.ws[0][gdat.columns[gdat.table+4]+'2']='°C'
    gdat.ws[0].merge_cells(gdat.columns[gdat.table+1]+'3:'+gdat.columns[gdat.table+2]+'3')
    gdat.ws[0][gdat.columns[gdat.table+1]+'3']='min'
    gdat.ws[0].merge_cells(gdat.columns[gdat.table+1]+'4:'+gdat.columns[gdat.table+2]+'4')
    gdat.ws[0][gdat.columns[gdat.table+1]+'4']=gdat.QC.tTtR.min
    gdat.ws[0].merge_cells(gdat.columns[gdat.table+3]+'3:'+gdat.columns[gdat.table+4]+'3')
    gdat.ws[0][gdat.columns[gdat.table+3]+'3']='max'
    gdat.ws[0].merge_cells(gdat.columns[gdat.table+3]+'4:'+gdat.columns[gdat.table+4]+'4')
    gdat.ws[0][gdat.columns[gdat.table+3]+'4']=gdat.QC.tTtR.max
    
    gdat.ws[0].merge_cells(gdat.columns[gdat.table+5]+'2:'+gdat.columns[gdat.table+6]+'2')
    gdat.ws[0][gdat.columns[gdat.table+5]+'2']='Tmax [°C]'
    gdat.ws[0][gdat.columns[gdat.table+5]+'3']='min'
    gdat.ws[0][gdat.columns[gdat.table+5]+'4']=gdat.QC.Tmax.min
    gdat.ws[0][gdat.columns[gdat.table+6]+'3']='max'
    gdat.ws[0][gdat.columns[gdat.table+6]+'4']=gdat.QC.Tmax.max
    
    gdat.ws[0].merge_cells(gdat.columns[gdat.table+7]+'2:'+gdat.columns[gdat.table+8]+'2')
    gdat.ws[0][gdat.columns[gdat.table+7]+'2']='T60s [°C]'
    gdat.ws[0][gdat.columns[gdat.table+7]+'3']='min'
    gdat.ws[0][gdat.columns[gdat.table+7]+'4']=gdat.QC.T60.min
    gdat.ws[0][gdat.columns[gdat.table+8]+'3']='max'
    gdat.ws[0][gdat.columns[gdat.table+8]+'4']=gdat.QC.T60.max
    
    gdat.ws[0].merge_cells(gdat.columns[gdat.table+9]+'2:'+gdat.columns[gdat.table+10]+'2')
    gdat.ws[0][gdat.columns[gdat.table+9]+'2']='Imax [A]'
    gdat.ws[0][gdat.columns[gdat.table+9]+'3']='min'
    gdat.ws[0][gdat.columns[gdat.table+9]+'4']=gdat.QC.Imax.min
    gdat.ws[0][gdat.columns[gdat.table+10]+'3']='max'
    gdat.ws[0][gdat.columns[gdat.table+10]+'4']=gdat.QC.Imax.max
    
    gdat.ws[0].merge_cells(gdat.columns[gdat.table+11]+'2:'+gdat.columns[gdat.table+12]+'2')
    gdat.ws[0][gdat.columns[gdat.table+11]+'2']='I60s [A]'
    gdat.ws[0][gdat.columns[gdat.table+11]+'3']='min'
    gdat.ws[0][gdat.columns[gdat.table+11]+'4']=gdat.QC.I60.min
    gdat.ws[0][gdat.columns[gdat.table+12]+'3']='max'
    gdat.ws[0][gdat.columns[gdat.table+12]+'4']=gdat.QC.I60.max
    
    gdat.ws[0].merge_cells(gdat.columns[gdat.table+13]+'2:'+gdat.columns[gdat.table+14]+'2')
    gdat.ws[0][gdat.columns[gdat.table+13]+'2']='R [Ohm]'
    gdat.ws[0][gdat.columns[gdat.table+13]+'3']='min'
    gdat.ws[0][gdat.columns[gdat.table+13]+'4']=gdat.QC.R.min
    gdat.ws[0][gdat.columns[gdat.table+14]+'3']='max'
    gdat.ws[0][gdat.columns[gdat.table+14]+'4']=gdat.QC.R.max


    gdat.ws[0].column_dimensions[gdat.columns[gdat.table+1]].width = 1.71
    gdat.ws[0].column_dimensions[gdat.columns[gdat.table+2]].width = 2.5
    gdat.ws[0].column_dimensions[gdat.columns[gdat.table+3]].width = 2.5
    gdat.ws[0].column_dimensions[gdat.columns[gdat.table+4]].width = 2.71
    for i in range(10):
        gdat.ws[0].column_dimensions[gdat.columns[gdat.table+5+i]].width = 4.8


def con_for(gdat):
    red = PatternFill(start_color='FF5252',end_color='FF5252',fill_type='solid')
    green = PatternFill(start_color='64FF64',end_color='6DFF6D',fill_type='solid')
    yelow = PatternFill(start_color='ffff49',end_color='ffff49',fill_type='solid')
    neutral = PatternFill(start_color='EEEEEE',end_color='EEEEEE',fill_type='solid')
    cells2=[gdat.columns[gdat.table+1], gdat.columns[gdat.table+3], gdat.columns[gdat.table+5], gdat.columns[gdat.table+6], gdat.columns[gdat.table+7], gdat.columns[gdat.table+8], gdat.columns[gdat.table+9], gdat.columns[gdat.table+10], gdat.columns[gdat.table+11], gdat.columns[gdat.table+12], gdat.columns[gdat.table+13], gdat.columns[gdat.table+14]]
    #cells2=['L','N','P','Q','R','S','T','U','V','W','X','Y']
    
    gdat.ws[0].conditional_formatting.add(gdat.columns[0]+'2:'+gdat.columns[0]+str(gdat.length+1),CellIsRule(operator='equal', formula=['"OK"'], stopIfTrue=True, fill=green))
    #gdat.ws[0].conditional_formatting.add(gdat.columns[0]+'2:'+gdat.columns[0]+str(gdat.length+1),CellIsRule(operator='equal', formula=['ok'], stopIfTrue=True, fill=green))
    #gdat.ws[0].conditional_formatting.add(gdat.columns[0]+'2:'+gdat.columns[0]+str(gdat.length+1),CellIsRule(operator='equal', formula=['Ok'], stopIfTrue=True, fill=green))
    gdat.ws[0].conditional_formatting.add(gdat.columns[0]+'2:'+gdat.columns[0]+str(gdat.length+1),CellIsRule(operator='equal', formula=['"NOK"'], stopIfTrue=True, fill=red))
    #gdat.ws[0].conditional_formatting.add(gdat.columns[0]+'2:'+gdat.columns[0]+str(gdat.length+1),CellIsRule(operator='equal', formula=['NOk'], stopIfTrue=True, fill=red))
    #gdat.ws[0].conditional_formatting.add(gdat.columns[0]+'2:'+gdat.columns[0]+str(gdat.length+1),CellIsRule(operator='equal', formula=['Nok'], stopIfTrue=True, fill=red))
    #gdat.ws[0].conditional_formatting.add(gdat.columns[0]+'2:'+gdat.columns[0]+str(gdat.length+1),CellIsRule(operator='equal', formula=['nok'], stopIfTrue=True, fill=red))
    
    
    gdat.ws[0].conditional_formatting.add(gdat.columns[4]+'2:'+gdat.columns[gdat.table-2]+str(gdat.length+1),CellIsRule(operator='equal', formula=[0], stopIfTrue=True, fill=neutral))
    if gdat.R:
        col=6
    else:
        col=5
    for i in range(col):
        gdat.ws[0].conditional_formatting.add(gdat.columns[i+4]+'2:'+gdat.columns[i+4]+str(gdat.length+1),CellIsRule(operator='between', formula=[cells2[i*2]+'$4',cells2[i*2+1]+'$4'], stopIfTrue=True, fill=green))
        gdat.ws[0].conditional_formatting.add(gdat.columns[i+4]+'2:'+gdat.columns[i+4]+str(gdat.length+1),CellIsRule(operator='notBetween', formula=[cells2[i*2]+'$4',cells2[i*2+1]+'$4'], stopIfTrue=True, fill=red))
    if gdat.TC:
        i += 5
        gdat.ws[0].conditional_formatting.add(gdat.columns[i]+'2:'+gdat.columns[i]+str(gdat.length+1),CellIsRule(operator='between', formula=[1050,1270], stopIfTrue=True, fill=green))
        gdat.ws[0].conditional_formatting.add(gdat.columns[i]+'2:'+gdat.columns[i]+str(gdat.length+1),CellIsRule(operator='between', formula=[1270,1280], stopIfTrue=True, fill=yelow))
        gdat.ws[0].conditional_formatting.add(gdat.columns[i]+'2:'+gdat.columns[i]+str(gdat.length+1),CellIsRule(operator='notBetween', formula=[1050,1280], stopIfTrue=True, fill=red))
        
        gdat.ws[0].conditional_formatting.add(gdat.columns[i+1]+'2:'+gdat.columns[i+1]+str(gdat.length+1),CellIsRule(operator='between', formula=[30,60], stopIfTrue=True, fill=green))
        gdat.ws[0].conditional_formatting.add(gdat.columns[i+1]+'2:'+gdat.columns[i+1]+str(gdat.length+1),CellIsRule(operator='between', formula=[20,30], stopIfTrue=True, fill=yelow))
        gdat.ws[0].conditional_formatting.add(gdat.columns[i+1]+'2:'+gdat.columns[i+1]+str(gdat.length+1),CellIsRule(operator='between', formula=[60,70], stopIfTrue=True, fill=yelow))
        gdat.ws[0].conditional_formatting.add(gdat.columns[i+1]+'2:'+gdat.columns[i+1]+str(gdat.length+1),CellIsRule(operator='notBetween', formula=[20,70], stopIfTrue=True, fill=red))
    return(gdat)
    
    
    
def check_col(col):
    col += 1
    if col<17:
        pass
    else:
        col=0
    return col
    
def write_graph(gdat, graph_nr=1, Curr=False):
    ml = messageLogger()
    vgraf=39
    gdat.ws.insert(1,gdat.wb.create_sheet('Grafi',1))
    colors=["000000", "26A320", "ff0000",  "1C1CC1", "F2FE14", "B300F5", "00ffd4", "ff8300", "ff0083", "cff00", "9b0000", "9b3900", "9b8400", "009b72", "450068", "680040", "659b00", "D30000", "999999"]
    col=0
    g_nr = 0
    sheets = len(gdat.ws)
    meas = len(gdat.filenames)
    tagi = 0
    for i in range(sheets-2):
        if (i+4) > sheets:
            iterj = meas - 10*i
        else:
            iterj = 10

        for j in range(iterj):
#            try:
                #xvalues = Reference(gdat.ws[i+2], min_col=j*5+1, min_row=3, max_row=6503)
                #values = Reference(gdat.ws[i+2], min_col=j*5+4, min_row=3, max_row=6503)
                #values1 = Reference(gdat.ws[i+2], min_col=j*5+2, min_row=3, max_row=6503)
                #values2 = Reference(gdat.ws[i+2], min_col=j*5+3, min_row=3, max_row=6503)
                
                if graph_nr<1:
                    ml.writeWarning('Število grafov nepravilno. Noben graf ni bil ustvarjen.', gdat)
                    break
                elif graph_nr == 1:
                    tag=''
                else:
                    tag=gdat.filenames[tagi]
                    tagi += 1
                
                ##values3 = Reference(gdat.ws[i+2], min_col=j*5+5, min_row=3, max_row=6503)
                
                #series = Series(values, xvalues,title='Temperature @ surface')
                #series1 = Series(values1, xvalues, title='Voltage')
                #series2 = Series(values2, xvalues, title='Current')
                
                #series.graphicalProperties.line.solidFill="D30000"
                #series1.graphicalProperties.line.solidFill="26A320"
                #series2.graphicalProperties.line.solidFill="1C1CC1"

                xvalues = Reference(gdat.ws[i+2], min_col=j*5+1, min_row=3, max_row=6503)
                values1 = Reference(gdat.ws[i+2], min_col=j*5+4, min_row=3, max_row=6503)
                
                
                
                    
                
                if g_nr == 0:
                    chart = ScatterChart()
                    chart1 = ScatterChart()
                    chart1.y_axis.axId = 200
                    chart1.title = gdat.filenames[i*10+j].strip('.txt')
                    
                    #title = gdat.filenames[i*10+j].strip('.txt')
                    cpt = CharacterProperties(sz=int(gdat.graphSett.tfG*100))
                    #print('Velikost naslova:', gdat.graphSett.tfG)
                    #tPara = [Paragraph(pPr=ParagraphProperties(defRPr=cpt), r=RegularTextRun(t=s)) for s in title.split("\n")]
                    #chart1.title.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cpt), endParaRPr=cpt)])
                    chart1.title.tx.rich.p[0].r[0].rPr = cpt#RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cpt), endParaRPr=cpt)])
                    #chart1.title = gdat.filenames[i*10+j].strip('.txt')
                    chart1.style = 13
                    chart1.x_axis.title = 't [s]'
                    chart.y_axis.title = 'T [°C]'
                    chart1.y_axis.title = 'U [V], I [A]'
                    chart.y_axis.majorGridlines = None
                    
                    ###
                    cp = CharacterProperties(sz=int(gdat.graphSett.atfG*100))
                    #xtStr = u"t [s]"
                    #ytStr = u"T [°C]"
                    #y1tStr = u"U [V], I [A]"
                    
                    #myChart.x_axis.title = ""
                    #myChart.y_axis.title = ""
                    #xPara = [Paragraph(pPr=ParagraphProperties(defRPr=cp), r=RegularTextRun(t=s)) for s in xtStr.split("\n")]
                    #yPara = [Paragraph(pPr=ParagraphProperties(defRPr=cp), r=RegularTextRun(t=s)) for s in ytStr.split("\n")]
                    #y1Para = [Paragraph(pPr=ParagraphProperties(defRPr=cp), r=RegularTextRun(t=s)) for s in y1tStr.split("\n")]
                    #chart1.x_axis.title.tx.rich.paragraphs = xPara
                    #chart.y_axis.title.tx.rich.paragraphs = yPara
                    #chart1.y_axis.title.tx.rich.paragraphs = y1Para
                    chart1.x_axis.title.tx.rich.p[0].r[0].rPr = cp
                    chart1.y_axis.title.tx.rich.p[0].r[0].rPr = cp
                    chart.y_axis.title.tx.rich.p[0].r[0].rPr = cp
                    cp = CharacterProperties(sz=int(gdat.graphSett.afG*100))
                    chart1.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                    chart1.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                    chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                    ###
                    
                    chart1.height = 20
                    chart1.width = 40
                    
                    values = Reference(gdat.ws[i+2], min_col=j*5+2, min_row=3, max_row=6503)
                    series = Series(values, xvalues, title='Voltage')
                    series.graphicalProperties.line.solidFill=colors[col]
                    chart1.series.append(series)
                    col += 1
                    chart1.layout=Layout(
                        manualLayout=ManualLayout(
                            x=0.08, y=-0.05,
                            h=0.85, w=0.9,
                        )
                    )

                    chart1.legend.layout = Layout(
                        manualLayout=ManualLayout(
                            yMode='edge',
                            xMode='edge',
                            x=0.5, y=0.9,
                            h=0.1, w=0.5
                        ))
                    
                cp = CharacterProperties(sz=int(gdat.graphSett.lfG*100))               
                chart1.legend.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
                    
                    
                series1 = Series(values1, xvalues,title='Temperature @ surface '+tag) 
                series1.graphicalProperties.line.solidFill=colors[col]
                col= check_col(col)
                
                
                
                if g_nr==0 and not Curr:
                    values2 = Reference(gdat.ws[i+2], min_col=j*5+3, min_row=3, max_row=6503)
                    series2 = Series(values2, xvalues, title='Current')
                    series2.graphicalProperties.line.solidFill=colors[col]
                    col= check_col(col)
                    chart1.series.append(series2)
                elif Curr:
                    values2 = Reference(gdat.ws[i+2], min_col=j*5+3, min_row=3, max_row=6503)
                    series2 = Series(values2, xvalues, title='Current '+tag)
                    series2.graphicalProperties.line.solidFill=colors[col]
                    col= check_col(col)
                    chart1.series.append(series2)
               
                
                

                
                if gdat.TC:
                    values3 = Reference(gdat.ws[i+2], min_col=j*5+5, min_row=3, max_row=6503)
                    series3 = Series(values3, xvalues,title='Temperature @ inside '+tag)
                    series3.graphicalProperties.line.solidFill=colors[col]
                    col= check_col(col)
                    chart.series.append(series3)
                
                chart.series.append(series1)
                
                

                g_nr += 1
                if g_nr>graph_nr-1:
                    chart.y_axis.crosses = "max"
                    chart1.x_axis.scaling.min = gdat.graphSett.talG[0]
                    chart.y_axis.scaling.min = gdat.graphSett.TalG[0]
                    chart1.x_axis.scaling.max = gdat.graphSett.talG[1]+round(gdat.MeasStart[i*10+j]/100)#+gdat.MeasDur
                    chart.y_axis.scaling.max = gdat.graphSett.TalG[1]
                    chart1.y_axis.scaling.min = gdat.graphSett.UIalG[0]
                    chart1.y_axis.scaling.max = gdat.graphSett.UIalG[1]
                    
                    chart1 += chart
                    #chart += chart1
                    gdat.ws[1].add_chart(chart1, "A"+str(round(i*10/graph_nr*vgraf+((j+1)/graph_nr-1)*vgraf+1)))
                    g_nr=0
                    col=0
                    chart = ScatterChart()
                    chart1 = ScatterChart()
                    chart1.y_axis.axId = 200
                    #chart1.title = ""#u""
                    #title = 'lalal'#gdat.filenames[i*10+j].strip('.txt')
                    #cpt = CharacterProperties(sz=int(gdat.graphSett.tfG*100))
                    #tPara = [Paragraph(pPr=ParagraphProperties(defRPr=cpt), r=RegularTextRun(t=s)) for s in title.split("\n")]
                    #chart1.title.tx.rich.paragraphs = tPara
                    #chart1.title = gdat.filenames[i*10+j].strip('.txt')
                    chart1.style = 13
#                    chart1.x_axis.title = 't [s]'
#                    chart.y_axis.title = 'T [°C]'
#                    chart1.y_axis.title = 'U [V], I [A]'
                    
                    chart.y_axis.majorGridlines = None
#                    chart1.y_axis.minorGridlines = 2.5

                    chart1.height = 20
                    chart1.width = 40
#            except:
#                chart.y_axis.crosses = "max"
#                chart1.x_axis.scaling.min = 0
#                chart.y_axis.scaling.min = 700
#                chart1.x_axis.scaling.max = round(gdat.MeasStart[i*10+j-1]/100)+gdat.MeasDur
#                chart.y_axis.scaling.max = 1300
#                chart1.y_axis.scaling.min = 0
#                chart1.y_axis.scaling.max = 30
#                chart1 += chart
#                #chart += chart1
#                gdat.ws[1].add_chart(chart1, "A"+str(round(i*10/graph_nr*vgraf+j/graph_nr*vgraf+1)))
#                break
#    c2.series.append(series2)
    return gdat
    
    
    
def define_TC(gdat):
    if gdat.ws[1][gdat.locations[0][5]+str(gdat.MeasStart[0]+1000)].value>900:
        return True
    else:
        return False
        
        
        
def table_head(gdat):
    labels=['OK/NOK', 'Oznaka podatkov', 'Žig svečke', 'Oznaka svečke', 't'+str(gdat.QC.TtR)+'C [s]', 'Tmax [°C]', 'T60s [°C]', 'Imax [A]', 'I60s [A]']
    if gdat.R:
        labels.append('R [Ω]')
    if gdat.gdat.nastavitveA.TT:
        labels.append('Odpoved')
    if gdat.TC:
        labels.append('Tmax(TC) [°C]')
        labels.append('T60(TC)-T60 [°C]')
    for i in range(len(labels)):
        gdat.ws[0][gdat.columns[i]+'1']=labels[i]
    color=PatternFill(patternType='solid', fill_type='solid', fgColor=Color('e0e0e0'))
    for j in range(4):
        for i in range(gdat.length+1):
            if j==3:
                color=PatternFill(patternType='solid', fill_type='solid', fgColor=Color('c1c1c1'))
            if j<3:
                gdat.ws[0].column_dimensions[gdat.columns[j]].width = 8
            gdat.ws[0][gdat.columns[j]+str(i+1)].fill = color
    for i in range(gdat.table-5):
        gdat.ws[0][gdat.columns[i+4]+'1'].fill = color
    if not gdat.mark:
        gdat.ws[0].column_dimensions.group('C', hidden=True)
    return gdat
    
    




def table_data(gdat,line,wsh,tMer,tcol,Tcol,Ucol,Icol,TCcol,TCc, worksheet):
    try:
        gdat.ws[0][gdat.columns[3]+str(line)]=int(gdat.filenames[line-2].strip('.txt'))
    except:
        gdat.ws[0][gdat.columns[3]+str(line)]=gdat.filenames[line-2].strip('.txt')
    i = 1
    while True:
        i += 1
        try:
            t60=str(round((tMer+gdat.ws[worksheet][tcol+str(gdat.MeasStart[-1]+i)].value)*100))
            break
        except:
            pass
    gdat.ws[0][gdat.columns[1]+str(line)]=line-1
    gdat.ws[0][gdat.columns[5]+str(line)]='=MAX('+wsh+Tcol+':'+Tcol+')'
    gdat.ws[0][gdat.columns[6]+str(line)]='='+wsh+Tcol+t60
    gdat.ws[0][gdat.columns[7]+str(line)]='=MAX('+wsh+Icol+':'+Icol+')'
    gdat.ws[0][gdat.columns[8]+str(line)]='='+wsh+Icol+t60
    gdat.locations.append([wsh,tcol,Ucol,Icol,Tcol,TCcol])
    if gdat.TC:
        gdat.ws[0][gdat.columns[9+TCc]+str(line)]='=MAX('+wsh+TCcol+':'+TCcol+')'
        gdat.ws[0][gdat.columns[10+TCc]+str(line)]='='+wsh+TCcol+t60+'-'+wsh+Tcol+t60
    return gdat
    

    
    
    
    
def write_table(gdat):
    ml = messageLogger()
    TCcol='A' ########################################pazi!!!!!
    ml.writeProgressInfo('Ustvarjam tabelo s parametri...', gdat)
    header = headers()
    gdat.ws.insert(0, gdat.wb.create_sheet('Tabela',0))
    TCc=0
    if gdat.R:
        TCc += 1
    if gdat.gdat.nastavitveA.TT:
        TCc += 1
    line=1
    dat=-9
    flag=0
    gdat.MeasStart = de.findMeasStart(gdat.data)
    for worksheet in range(len(gdat.ws)):
        maxcol=gdat.maxlines
        TtR=gdat.QC.TtR
        tMer=gdat.MeasDur
        wsh="'"+str(dat)+" do "+str(dat+9)+"'!"
        if dat>0:
            for column in range(len(gdat.columns)):
                #print(column)
                #print('headers:',gdat.ws[worksheet][gdat.columns[column]+'2'].value, header.T)
                if gdat.ws[worksheet][gdat.columns[column]+'2'].value in header.t:#=='t[s]':
                    if column==0:
                        pass
                    else:
                        gdat=table_data(gdat,line,wsh,tMer,tcol,Tcol,Ucol,Icol,TCcol,TCc, worksheet) #def table_data(gdat,line,wsh,tMer,tcol,Tcol,Ucol,Icol,TCcol,TCc)
                        #gdat.ws[0][gdat.columns[maxT]+str(line)]="=LOOKUP(N2+1;'1-10'!D$3:D1003;'1-10'!A$3:A$1003)"
                        #print(gdat.columns[maxT]+str(line),"=LOOKUP(N2+1;"+wsh+Tcol+"$3:"+Tcol+"1003;"+wsh+tcol+"$3:"+tcol+"$1003)-AB3")
                    line+=1
                    flag=0
                    tcol=gdat.columns[column]
                elif gdat.ws[worksheet][gdat.columns[column]+'2'].value in header.T:
                    Tcol=gdat.columns[column]
                    flag+=1
#                    for i in range(maxcol):
#                        if gdat.ws[worksheet][Tcol+str(3+i)].value>TtR:
#                            gdat.ws[0][gdat.columns[2]+str(line)]='='+wsh+tcol+str(i)
#                            break
                elif gdat.ws[worksheet][gdat.columns[column]+'2'].value in header.TC:#'TC-K[°C]' or gdat.ws[worksheet][gdat.columns[column]+'2'].value=='TC[C]' or gdat.ws[worksheet][gdat.columns[column]+'2'].value=='TC[°C]':
                    TCcol=gdat.columns[column]
                    
                elif gdat.ws[worksheet][gdat.columns[column]+'2'].value in header.U:#=='U[V]':
                    Ucol=gdat.columns[column]
                    for j in range(maxcol-3):
                        #print('U: ',gdat.ws[worksheet][Ucol+str(j+3)].value)
                        if gdat.ws[worksheet][Ucol+str(j+3)].value>1:
                            #MesStart=j
                            if j < 1:
                                j = 1
                                #print('!!!Warning!!! Not typical measurement detected - Early voltage start in measurement', gdat.filenames[len(gdat.MeasStart)])
                            #gdat.MeasStart.append(j)
                            #print('mstN:',j)
                            flag+=1
                            break
                    #if j>maxcol-10:
                        #print('!!!Warning!!! Not typical measurement detected -  '+ gdat.filenames[len(gdat.MeasStart)])
                elif gdat.ws[worksheet][gdat.columns[column]+'2'].value in header.I:#=='I[A]':
                    Icol=gdat.columns[column]
                if flag>1:
                    for i in range(maxcol):
                        #if i<50:
                            #print('T: ',Tcol+str(3+i))
                        try:
                            if gdat.ws[worksheet][Tcol+str(3+i)].value>TtR:
                                gdat.ws[0][gdat.columns[4]+str(line)]='='+wsh+tcol+str(i+3)+'-'+wsh+tcol+str(gdat.MeasStart[-1]+2)
                                break
                        except:
                            gdat.ws[0][gdat.columns[4]+str(line)]='='+wsh+tcol+str(i+2)+'-'+wsh+tcol+str(gdat.MeasStart[-1]+2)
                            break
                        
            gdat=table_data(gdat,line,wsh,tMer,tcol,Tcol,Ucol,Icol,TCcol,TCc, worksheet) #def table_data(gdat,line,wsh,tMer,tcol,Tcol,Ucol,Icol,TCcol,TCc)
            #gdat.ws[0][gdat.columns[maxT]+str(line)]="=LOOKUP(N2+1;"+wsh+Tcol+"$3:"+Tcol+"1003;"+wsh+tcol+"$3:"+tcol+"$1003)-AB3"
        dat+=10
    
    write_QC(gdat)
    gdat=table_head(gdat)
    gdat=con_for(gdat)
    gdat=write_graph(gdat, graph_nr=gdat.graph, Curr=gdat.curr)
    ml.writeProgressInfo('Tabela s parametri ustvarjena.', gdat)
    return gdat
 
#def create_TC_header(gdat, wb, nr):
#    print('v TC headru')
#    wb['Poročilo'].add_image(Image(gdat.path+'Graphic\Pic\Hidria_AET.png'), 'A1')
#    wb['Poročilo']['A2']='Date: '+ gdat.TC_header.date
#    wb['Poročilo']['E2']='Order No./Customer: '+gdat.TC_header.order_nr
#    wb['Poročilo']['B9']=gdat.TC_header.production_date
#    wb['Poročilo']['B11']=nr
#    wb['Poročilo']['F7']=gdat.TC_header.gp_nr
#    wb['Poročilo']['F9']=gdat.TC_header.nominal_v
#    if gdat.TC_header.nominal_v == '5V':   
#        wb['Poročilo']['A19']= '11V-2s'
#        wb['Poročilo']['A21']= '5V-58s'
#        print('5V')
#    elif gdat.TC_header.nominal_v == '4,4V':   
#        wb['Poročilo']['A19']= '11V-1,9s'
#        wb['Poročilo']['A21']= '4,4V-58,1s'
#        print('4,4V')
#    elif gdat.TC_header.nominal_v == '11V':   
#        wb['Poročilo']['A19']= '11V-58s'
#        wb['Poročilo']['A21']= ' '
#        print('11V')
#    else:
#        raise RuntimeError('unknown nominal voltage')
#    wb['Poročilo']['E45']='Prepared by/signature: ' + gdat.TC_header.creator
#    wb['Poročilo'].add_image(Image(gdat.path+'Graphic\Pic\Legend.png'), 'B49')
#    
#    
#    wb['Poročilo'].column_dimensions['F'].width = 9.10
#    wb['Poročilo'].column_dimensions['G'].width = 9.10
#    wb['Poročilo'].column_dimensions['H'].width = 9.10
#    wb['Poročilo'].column_dimensions['I'].width = 9.10
#    wb['Poročilo'].column_dimensions['J'].width = 9.10
#    return wb
    
    
def create_TC_report(gdat, first_TC=1):
    label = gdat.folder.split('/')[-1].split('\\')[-1]
    TCwb = load_workbook(gdat.path+'Templates\TC_template.xlsx')
    #print('a:', TCwb)
#    if gdat.folder == 'measurements':
#        path = ''
#    else:
#        path = gdat.path
    Anwb = load_workbook(gdat.folder + '\\' + label +'AFM.xlsx') ####\\
    Aws = Anwb['Tabela']

    if gdat.folder == 'measurements':
        path = ''
    else:
        path = gdat.folder + '\\'
    dataTC=rd.measurment_data(gdat.gdat, TC=False,MaxLines=27002,MeasDur=270, folder=gdat.folder + '/TC800-1100')
    dataTC.filenames=dm.rearrange_up(rd.get_data_names(path + 'TC800-1100'))
    dataTC.length = len(dataTC.filenames)
    dataTC.read_data()
    dataTC=dm.transform_to_write(dataTC)
    k=0
    lastFile = []
    for i in range(gdat.length):
        #print('i: ', i)
        #tempImg = 'Graphic/Pic/temp.png'
        if Aws['A'+str(i+2)].value=='OK' or Aws['A'+str(i+2)].value=='Ok' or Aws['A'+str(i+2)].value=='ok':
            for j in range(27002):
                for l in range(5):
                    try:
                        TCwb['Meritev 11V-2s 5V-58s'][gdat.columns[l]+str(j+1)]=gdat.data[i][j][l]
                    except:
                        pass
                    #try:
                    TCwb['Meritve 800-1100'][gdat.columns[l]+str(j+1)]=dataTC.data[k][j][l]
                    #except:
                        #print('j:',j)
                        #break
            #print('b:', TCwb)
            TCwb = TCrm.write_TC_header(gdat,TCwb,k+first_TC)
            #print('c:', TCwb)
            TCwb = TCrm.write_TC_diff_graph(TCwb)
            #print('d:', TCwb)
            TCwb = TCrm.write_TC_func_graph(TCwb,gdat.MeasStart[i])
            #print('e:', TCwb)
            TCwb = TCrm.write_800_1100_graph(TCwb, gdat.filenames[i], dataTC.filenames[k])
            #print('f:', TCwb)
            TCwb['Meritve 800-1100']['Z3']=round((gdat.MeasStart[i]-1)/100)
            #print('g:', TCwb)
            fp = path + 'TCRTG\\' + gdat.filenames[i] + '.jpg'
            im = PILImage.open(fp)
            im1 = im.crop((0, 0, 1000, 667))
            im1.load()
            #im = im.resize((173, 130), PILImage.ANTIALIAS)
            memfile = BytesIO()
            im1.save(memfile, format='png')
            #im.save(fp=tempImg, format='png')
            #TCwb['Poročilo'].add_image(Image(tempImg), 'D75')#, size=(170, 305)
            img = Image(memfile)
            img.width = 195
            img.height = 130
            TCwb['Poročilo'].add_image(img, 'D75')
            #print('h:', TCwb)
            #wb['Poročilo'].add_image(Image('TCRTG\\a.jpg', size=(170, 305)), 'D75')
            lastFile.append(save(gdat, TCwb, 'Instrumented glow plug data sheet_'+str(k+first_TC)+'_'+gdat.filenames[i]+'.xlsx'))
            k += 1
    return lastFile



class headers():
    
    def __init__(self):
        self.readHeaders()
        
    def readHeaders(self):
        headerFile = open('settings/headers.txt', 'r', encoding="latin-1")
        lines = headerFile.readlines()
        headerFile.close()
        for line in lines:
            line = line.strip('\n').split('\t')
            if line[0] == 'T':
                self.T = line[1:]
            elif line[0] == 'U':
                self.U = line[1:]
            elif line[0] == 'I':
                self.I = line[1:]
            elif line[0] == 'TC':
                self.TC = line[1:]
            elif line[0] == 't':
                self.t = line[1:]